import openpyxl, os, pprint
from openpyxl.utils.cell import get_column_letter, column_index_from_string
os.chdir('D:/Drive/Code/ATBSWP/Chapter_12')
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
columnB = list(list(sheet.columns)[1])
for cell in columnB:
    print(cell.value)

# sheet = wb.get_sheet_by_name('Sheet1')

# pprint.pprint(list(sheet['A1':'C3']))
# for row in sheet.iter_rows('B%s:B%s' %(sheet.min_row,sheet.max_row)):
#     for cell in row:
#         print(cell.value)

print('....')


input()
